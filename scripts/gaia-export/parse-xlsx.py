#!/usr/bin/env python3
"""Parse the Member Profiles xlsx export into accounts-xlsx.json.

The July 2026 spreadsheet export is richer than the original scrape
(accounts.json): per pokemon it carries explicit gender, shiny, level,
friendship, shadow, purification, nickname and held item, and items are
grouped per character. This script converts it into a JSON overlay that
functions/scripts/upload-gaia-exports.mjs merges over accounts.json.

Usage: python3 parse-xlsx.py /path/to/Team_Snagem_Member_Profiles_Export.xlsx
Writes accounts-xlsx.json next to this script.
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent

# Currency rows hide inside per-character item lists.
CURRENCY_ROWS = {
    "snag coins": "snagCoins",
    "snag coin": "snagCoins",
    "snag emblem": "snagEmblems",
    "snag emblems": "snagEmblems",
    "snag emblem piece": "emblemPieces",
    "snag emblem pieces": "emblemPieces",
    "gengar tokens": "gengarTokens",
    "gengar token": "gengarTokens",
    "gengar coins": "gengarTokens",
}

FORM_PREFIXES = ("Alolan", "Galarian", "Hisuian", "Paldean", "Mega", "Bloodmoon")


def cell(v):
    return "" if v is None else str(v).strip()


def parse_sheet(ws):
    account = {
        "gaiaName": "",
        "threads": [],
        "characters": [],
        "pokemon": [],
        "items": [],
        "currency": {"snagCoins": 0, "snagEmblems": 0, "emblemPieces": 0, "gengarTokens": 0},
    }
    mode = None  # None | "pokemon" | "items"
    character = ""
    for row in ws.iter_rows(values_only=True):
        first = cell(row[0] if row else None)
        if not first:
            continue
        if first.startswith("Account:"):
            account["gaiaName"] = first.split(":", 1)[1].strip()
            continue
        if first.startswith("Thread:"):
            account["threads"].append(first[len("Thread:"):].strip())
            continue
        if first.startswith("Character:"):
            character = first.split(":", 1)[1].strip()
            account["characters"].append(character)
            mode = None
            continue
        if first == "Species":
            mode = "pokemon"
            continue
        if first == "Item Name":
            mode = "items"
            continue
        if first.startswith("(no items"):
            mode = None
            continue
        if mode == "pokemon":
            raw_species = first
            species = raw_species
            form = ""
            for pref in FORM_PREFIXES:
                if species.startswith(pref + " "):
                    form = pref
                    species = species[len(pref) + 1:]
                    break

            def col(i):
                return cell(row[i]) if len(row) > i else ""

            def num(i):
                s = col(i)
                try:
                    return int(float(s))
                except ValueError:
                    return 0

            level = max(1, min(100, num(3) or 1))
            entry = {
                "raw": raw_species,
                "species": species,
                "form": form,
                "gender": col(1).upper() if col(1).upper() in ("M", "F") else "",
                "shiny": col(2).upper().startswith("Y"),
                "level": level,
                "friendship": max(0, min(100, num(4))),
                "shadow": max(0, min(100, num(5))),
                "purification": max(0, min(100, num(6))),
                "nickname": col(7),
                "heldItem": col(8),
                "levelSource": col(9),
                "note": col(10),
                "character": character,
            }
            account["pokemon"].append(entry)
            continue
        if mode == "items":
            qty_s = cell(row[1]) if len(row) > 1 else ""
            try:
                qty = max(1, int(float(qty_s)))
            except ValueError:
                qty = 1
            key = CURRENCY_ROWS.get(first.lower())
            if key:
                account["currency"][key] += qty
                continue
            account["items"].append(
                {
                    "raw": first,
                    "qty": qty,
                    "note": cell(row[2]) if len(row) > 2 else "",
                    "character": character,
                }
            )
    return account


def main():
    if len(sys.argv) != 2:
        sys.exit("usage: parse-xlsx.py <export.xlsx>")
    wb = load_workbook(sys.argv[1], read_only=True, data_only=True)
    accounts = []
    for name in wb.sheetnames:
        if name == "Index":
            continue
        acc = parse_sheet(wb[name])
        if not acc["gaiaName"]:
            acc["gaiaName"] = name
        accounts.append(acc)
    out = HERE / "accounts-xlsx.json"
    out.write_text(json.dumps(accounts, indent=1, ensure_ascii=False))
    total_p = sum(len(a["pokemon"]) for a in accounts)
    total_i = sum(len(a["items"]) for a in accounts)
    total_c = sum(len(a["characters"]) for a in accounts)
    print(f"{len(accounts)} accounts | {total_c} characters | {total_p} pokemon | {total_i} item stacks")
    coins = sum(a["currency"]["snagCoins"] for a in accounts)
    print(f"currency totals: coins {coins}, emblems {sum(a['currency']['snagEmblems'] for a in accounts)}, pieces {sum(a['currency']['emblemPieces'] for a in accounts)}, gengar {sum(a['currency']['gengarTokens'] for a in accounts)}")


if __name__ == "__main__":
    main()
